#!/usr/bin/env python3
"""
benefits_tracker.py, builds the FutureHire benefits-tracker spreadsheet from the
time-savings model in docs/before_after_comparison.md.

Illustrative model. All figures are scenario estimates under the assumptions below,
not measured client outcomes. Edit the ASSUMPTIONS block and re-run to update the
workbook, which includes a live prioritisation-matrix sheet and an adoption-log
template for recording real pilot measurements.

Usage:
    pip install openpyxl
    python tools/benefits_tracker.py    # -> deliverables/FutureHire_Benefits_Tracker.xlsx
"""

import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl")

# ----------------------------------------------------------------------------
# ASSUMPTIONS  (edit these, everything else recomputes)
# ----------------------------------------------------------------------------
RECRUITERS = 24
WEEKS_PER_YEAR = 46
RATE = 75  # AUD/hour blended recruiter cost of time

PILOTS = [
    # name, before_min, after_min, per_week_volume
    ("Job-ad drafting",        45, 18, 25),
    ("Interview summarisation", 20,  8, 90),
    ("Hiring-manager updates",  15,  6, 80),
]

# Prioritisation matrix (value, candidate-impact, feasibility, risk)
OPPORTUNITIES = [
    ("Hiring-manager updates", 4, 3, 5, 1),
    ("Job-ad drafting",        4, 4, 5, 2),
    ("Interview summarisation",5, 4, 4, 2),
    ("Candidate email drafting",4,4, 5, 2),
    ("Scheduling & follow-up", 3, 3, 3, 2),
    ("Bullhorn auto-population",3,2, 2, 3),
]

# ----------------------------------------------------------------------------
# Styling
# ----------------------------------------------------------------------------
NAVY, TEAL = "0F1B2A", "14655C"
HFILL = PatternFill("solid", fgColor=NAVY)
SUBFILL = PatternFill("solid", fgColor=TEAL)
INPUT_FILL = PatternFill("solid", fgColor="F5EDE0")
GOOD_FILL = PatternFill("solid", fgColor="E6F0EA")
HFONT = Font(color="FFFFFF", bold=True)
TITLE = Font(name="Cambria", color=NAVY, bold=True, size=16)
BOLD = Font(bold=True)
ITAL = Font(italic=True, color="666666")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hrow(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=h)
        c.fill = HFILL; c.font = HFONT; c.border = BORDER


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def compute():
    rows = []
    tot_hrs = tot_val = 0
    for name, before, after, vol in PILOTS:
        saved = before - after
        red = saved / before
        hrs_yr = vol * saved / 60 * WEEKS_PER_YEAR
        val = hrs_yr * RATE
        tot_hrs += hrs_yr; tot_val += val
        rows.append(dict(name=name, before=before, after=after, saved=saved,
                         red=red, vol=vol, hrs=hrs_yr, val=val))
    return rows, tot_hrs, tot_val


def build(path):
    rows, tot_hrs, tot_val = compute()
    wb = Workbook()

    # Summary
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "FutureHire Recruitment AI, Benefits Tracker"; ws["A1"].font = TITLE
    ws["A2"] = "Illustrative model, figures are scenario estimates under the assumptions shown."
    ws["A2"].font = ITAL
    tiles = [("Three-pilot hours / yr", f"{tot_hrs:,.0f} hrs"),
             ("Indicative value / yr", f"AUD ${tot_val:,.0f}"),
             ("Per recruiter / week", f"{tot_hrs/RECRUITERS/WEEKS_PER_YEAR:.1f} hrs"),
             ("Min. task-time cut", f"{min(r['red'] for r in rows)*100:.0f}%")]
    row = 4
    for i, (label, val) in enumerate(tiles):
        col = 1 + i*2
        c = ws.cell(row=row, column=col, value=label); c.fill = SUBFILL; c.font = HFONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        v = ws.cell(row=row+1, column=col, value=val); v.font = Font(size=14, bold=True, color=NAVY)
        v.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions[get_column_letter(col+1)].width = 3
    ws["A8"] = ("Budget benchmark: engagement fee AUD $22,000. Modelled first-year value of the three "
                "pilots ≈ AUD ${:,.0f} (~{:.0f}× the fee) on the stated scenario assumptions. Targets: "
                "≥40% admin-time cut and 50% candidate-summary-prep cut, both cleared. Primary metric is "
                "task-time %, robust to the volume/$ assumptions.").format(tot_val, tot_val/22000)
    ws["A8"].alignment = Alignment(wrap_text=True); ws.merge_cells("A8:H11")

    # Assumptions
    wsa = wb.create_sheet("Assumptions")
    wsa["A1"] = "Assumptions (editable inputs)"; wsa["A1"].font = TITLE
    wsa["A2"] = "Highlighted cells are inputs. Change them, re-run the script."; wsa["A2"].font = ITAL
    hrow(wsa, 4, ["Parameter", "Value", "Unit / note"])
    base = [("Recruiters", RECRUITERS, "people"),
            ("Working weeks / year", WEEKS_PER_YEAR, "weeks"),
            ("Blended recruiter cost", RATE, "AUD / hour")]
    rr = 5
    for n, v, u in base:
        wsa.cell(row=rr, column=1, value=n)
        c = wsa.cell(row=rr, column=2, value=v); c.fill = INPUT_FILL; c.border = BORDER
        wsa.cell(row=rr, column=3, value=u); rr += 1
    rr += 1
    wsa.cell(row=rr, column=1, value="Per-pilot inputs").font = BOLD; rr += 1
    hrow(wsa, rr, ["Pilot", "Before (min)", "After (min)", "Volume / week"]); rr += 1
    for name, before, after, vol in PILOTS:
        wsa.cell(row=rr, column=1, value=name)
        for col, val in [(2, before), (3, after), (4, vol)]:
            c = wsa.cell(row=rr, column=col, value=val); c.fill = INPUT_FILL; c.border = BORDER
        rr += 1
    widths(wsa, [30, 14, 14, 16])

    # Pilot models
    wsp = wb.create_sheet("Pilot models")
    wsp["A1"] = "Pilot models, computed from Assumptions"; wsp["A1"].font = TITLE
    hrow(wsp, 3, ["Pilot", "Before", "After", "Reduction", "≥ target",
                  "Volume/wk", "Hours/yr", "Value/yr (AUD)"])
    rr = 4
    for r in rows:
        wsp.cell(row=rr, column=1, value=r["name"])
        wsp.cell(row=rr, column=2, value=f'{r["before"]} min')
        wsp.cell(row=rr, column=3, value=f'{r["after"]} min')
        wsp.cell(row=rr, column=4, value=f'{r["red"]*100:.0f}%')
        tgt = wsp.cell(row=rr, column=5, value="Yes" if r["red"] >= 0.40 else "No")
        if r["red"] >= 0.40: tgt.fill = GOOD_FILL
        wsp.cell(row=rr, column=6, value=r["vol"])
        wsp.cell(row=rr, column=7, value=round(r["hrs"]))
        wsp.cell(row=rr, column=8, value=f'${r["val"]:,.0f}')
        for cc in range(1, 9): wsp.cell(row=rr, column=cc).border = BORDER
        rr += 1
    wsp.cell(row=rr, column=1, value="TOTAL").font = BOLD
    wsp.cell(row=rr, column=7, value=round(tot_hrs)).font = BOLD
    wsp.cell(row=rr, column=8, value=f"${tot_val:,.0f}").font = BOLD
    wsp.cell(row=rr+2, column=1,
             value="Scenario estimates. Task-time % depends only on per-task time (measurable); "
                   "volume/$ scale value, not the %.").font = ITAL
    wsp.merge_cells(start_row=rr+2, start_column=1, end_row=rr+2, end_column=8)
    widths(wsp, [26, 10, 10, 11, 10, 11, 11, 15])

    # Prioritisation matrix
    wsm = wb.create_sheet("Prioritisation matrix")
    wsm["A1"] = "Use-case prioritisation (Priority = Value × Feasibility × Candidate-impact ÷ Risk)"
    wsm["A1"].font = TITLE
    hrow(wsm, 3, ["Use case", "Value", "Candidate impact", "Feasibility", "Risk", "Priority", "Decision"])
    scored = []
    for name, v, ci, f, risk in OPPORTUNITIES:
        pri = round(v * f * ci / risk)
        scored.append((name, v, ci, f, risk, pri))
    scored.sort(key=lambda x: -x[5])
    rr = 4
    for i, (name, v, ci, f, risk, pri) in enumerate(scored):
        dec = "Pilot" if name in ("Job-ad drafting", "Interview summarisation", "Hiring-manager updates") else f"Rank {i+1}"
        for col, val in [(1, name), (2, v), (3, ci), (4, f), (5, risk), (6, pri), (7, dec)]:
            c = wsm.cell(row=rr, column=col, value=val); c.border = BORDER
        if dec == "Pilot":
            for cc in range(1, 8): wsm.cell(row=rr, column=cc).fill = GOOD_FILL
        rr += 1
    wsm.cell(row=rr+1, column=1,
             value="Candidate screening/ranking is excluded by design (risk = 5): AI never decides on candidates.").font = ITAL
    wsm.merge_cells(start_row=rr+1, start_column=1, end_row=rr+1, end_column=7)
    widths(wsm, [26, 10, 16, 12, 8, 10, 12])

    # Adoption log
    wsl = wb.create_sheet("Adoption log")
    wsl["A1"] = "Adoption log, replace scenario estimates with measured pilot actuals"; wsl["A1"].font = TITLE
    wsl["A2"] = "AI Champions fill this in during the pilot and 90-day rollout."; wsl["A2"].font = ITAL
    hrow(wsl, 4, ["Week", "Workflow", "Items measured", "Avg time BEFORE",
                  "Avg time AFTER", "Actual reduction %", "Quality/fairness pass %", "Notes"])
    for i in range(12):
        for cc in range(1, 9): wsl.cell(row=5+i, column=cc).border = BORDER
        wsl.cell(row=5+i, column=1, value=i+1)
    widths(wsl, [8, 22, 14, 15, 15, 15, 18, 30])

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print("Wrote", path)
    for r in rows:
        print(f"  {r['name']}: {r['red']*100:.0f}% cut, {r['hrs']:,.0f} hrs/yr, ${r['val']:,.0f}")
    print(f"  TOTAL: {tot_hrs:,.0f} hrs/yr, AUD ${tot_val:,.0f} (~{tot_val/22000:.0f}x fee), "
          f"{tot_hrs/RECRUITERS/WEEKS_PER_YEAR:.1f} hrs/recruiter/week")


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    out = os.path.normpath(os.path.join(here, "..", "deliverables", "FutureHire_Benefits_Tracker.xlsx"))
    build(out)
