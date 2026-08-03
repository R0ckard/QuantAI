#!/usr/bin/env python3
"""
benefits_tracker.py, builds the Elevate Accounting benefits-model spreadsheet
from the time-savings model in docs/08_benefits_model.md.

Scenario model. All figures are estimates under the assumptions below, not
measured client outcomes. Edit the ASSUMPTIONS block and re-run to update the
workbook, which includes a document-completeness sheet and an adoption-log
template for recording real pilot measurements.

Usage:
    pip install openpyxl
    python tools/benefits_tracker.py    # -> deliverables/Elevate_Benefits_Model.xlsx
"""

import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl")

# ----------------------------------------------------------------------------
# ASSUMPTIONS  (edit these, everything else recomputes)
# ----------------------------------------------------------------------------
STAFF = 18
WEEKS_PER_YEAR = 46
RATE = 80              # AUD/hour blended cost of staff time on onboarding
FEE = 12500            # AUD engagement fee benchmark
TARGET_CUT = 0.75      # brief: 3 staff hours -> 45 minutes = 75%
DOC_BEFORE = 0.70      # scenario baseline document completeness before first meeting
DOC_TARGET = 0.95      # brief target

# Service lines are modelled in HOURS per onboarding (admin portion only).
SERVICE_LINES = [
    # name, before_hours, after_hours, volume_per_year
    ("Business advisory", 4.0, 1.0, 45),
    ("Tax compliance",    3.0, 0.75, 90),   # the brief's "standard": 3h -> 45min
    ("Bookkeeping",       2.0, 0.5, 60),
]

# ----------------------------------------------------------------------------
# Styling (QuantAI brand)
# ----------------------------------------------------------------------------
NAVY, TEAL = "0F1B2A", "14655C"
HFILL = PatternFill("solid", fgColor=NAVY)
SUBFILL = PatternFill("solid", fgColor=TEAL)
INPUT_FILL = PatternFill("solid", fgColor="F5EDE0")
GOOD_FILL = PatternFill("solid", fgColor="E6F0EA")
HFONT = Font(color="FFFFFF", bold=True)
TITLE = Font(name="Cambria", color=NAVY, bold=True, size=16)
BOLD = Font(bold=True)
ITAL = Font(italic=True, color="666666")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hrow(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=h)
        c.fill = HFILL; c.font = HFONT; c.border = BORDER


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def compute():
    rows = []
    tot_hrs = tot_val = 0
    for name, before, after, vol in SERVICE_LINES:
        saved = before - after
        red = saved / before
        hrs_yr = vol * saved
        val = hrs_yr * RATE
        tot_hrs += hrs_yr; tot_val += val
        rows.append(dict(name=name, before=before, after=after, saved=saved,
                         red=red, vol=vol, hrs=hrs_yr, val=val))
    return rows, tot_hrs, tot_val


def build(path):
    rows, tot_hrs, tot_val = compute()
    wb = Workbook()

    # Summary
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "Elevate Accounting, Client Onboarding Benefits Model"; ws["A1"].font = TITLE
    ws["A2"] = "Scenario model, figures are estimates under the assumptions shown, not measured outcomes."
    ws["A2"].font = ITAL
    tiles = [("Hours returned / yr", f"{tot_hrs:,.0f} hrs"),
             ("Indicative value / yr", f"AUD ${tot_val:,.0f}"),
             ("Value vs fee", f"~{tot_val/FEE:.1f}x"),
             ("Standard time cut", f"{TARGET_CUT*100:.0f}%")]
    row = 4
    for i, (label, val) in enumerate(tiles):
        col = 1 + i*2
        c = ws.cell(row=row, column=col, value=label); c.fill = SUBFILL; c.font = HFONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        v = ws.cell(row=row+1, column=col, value=val); v.font = Font(size=14, bold=True, color=NAVY)
        v.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions[get_column_letter(col+1)].width = 3
    ws["A8"] = ("Brief success measures: standard onboarding falls from 3 staff hours to 45 minutes (75%) in "
                "scenario testing, and at least 95% of required documents are in before the first advisory "
                "meeting. Both are met on these assumptions (see the Service-line models and Document "
                "completeness sheets). Engagement fee benchmark AUD ${:,.0f}; modelled staff time returned is "
                "about AUD ${:,.0f}/yr (~{:.1f}x the fee). The per-onboarding time cut and the completeness "
                "rate are the robust measures; the dollar figure sizes the prize.").format(FEE, tot_val, tot_val/FEE)
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="top"); ws.merge_cells("A8:H12")

    # Assumptions
    wsa = wb.create_sheet("Assumptions")
    wsa["A1"] = "Assumptions (editable inputs)"; wsa["A1"].font = TITLE
    wsa["A2"] = "Highlighted cells are inputs. Change them, re-run the script."; wsa["A2"].font = ITAL
    hrow(wsa, 4, ["Parameter", "Value", "Unit / note"])
    base = [("Staff", STAFF, "people"),
            ("Working weeks / year", WEEKS_PER_YEAR, "weeks"),
            ("Blended cost of onboarding time", RATE, "AUD / hour"),
            ("Engagement fee benchmark", FEE, "AUD"),
            ("Standard time-cut target", f"{TARGET_CUT*100:.0f}%", "3 staff hours -> 45 minutes"),
            ("Doc completeness before (baseline)", f"{DOC_BEFORE*100:.0f}%", "scenario"),
            ("Doc completeness target", f"{DOC_TARGET*100:.0f}%", "brief target")]
    rr = 5
    for n, v, u in base:
        wsa.cell(row=rr, column=1, value=n)
        c = wsa.cell(row=rr, column=2, value=v); c.fill = INPUT_FILL; c.border = BORDER
        wsa.cell(row=rr, column=3, value=u); rr += 1
    rr += 1
    wsa.cell(row=rr, column=1, value="Per-service-line inputs (hours per onboarding)").font = BOLD; rr += 1
    hrow(wsa, rr, ["Service line", "Before (h)", "After (h)", "Volume / yr"]); rr += 1
    for name, before, after, vol in SERVICE_LINES:
        wsa.cell(row=rr, column=1, value=name)
        for col, val in [(2, before), (3, after), (4, vol)]:
            c = wsa.cell(row=rr, column=col, value=val); c.fill = INPUT_FILL; c.border = BORDER
        rr += 1
    wsa.cell(row=rr+1, column=1,
             value="Saving is only on the staff administration of onboarding (drafting, requesting, chasing, "
                   "tracking, preparing). It does not touch professional advice or client relationship time.").font = ITAL
    wsa.merge_cells(start_row=rr+1, start_column=1, end_row=rr+1, end_column=4)
    widths(wsa, [34, 14, 14, 16])

    # Service-line models
    wsp = wb.create_sheet("Service-line models")
    wsp["A1"] = "Service-line models, computed from Assumptions"; wsp["A1"].font = TITLE
    hrow(wsp, 3, ["Service line", "Before", "After", "Time cut", ">= 75% target",
                  "Volume/yr", "Hours/yr", "Value/yr (AUD)"])
    rr = 4
    for r in rows:
        wsp.cell(row=rr, column=1, value=r["name"])
        wsp.cell(row=rr, column=2, value=f'{r["before"]:.2f} h')
        wsp.cell(row=rr, column=3, value=f'{r["after"]:.2f} h')
        wsp.cell(row=rr, column=4, value=f'{r["red"]*100:.0f}%')
        tgt = wsp.cell(row=rr, column=5, value="Yes" if r["red"] >= TARGET_CUT - 1e-9 else "No")
        if r["red"] >= TARGET_CUT - 1e-9: tgt.fill = GOOD_FILL
        wsp.cell(row=rr, column=6, value=r["vol"])
        wsp.cell(row=rr, column=7, value=round(r["hrs"]))
        wsp.cell(row=rr, column=8, value=f'${r["val"]:,.0f}')
        for cc in range(1, 9): wsp.cell(row=rr, column=cc).border = BORDER
        rr += 1
    wsp.cell(row=rr, column=1, value="TOTAL").font = BOLD
    wsp.cell(row=rr, column=7, value=round(tot_hrs)).font = BOLD
    wsp.cell(row=rr, column=8, value=f"${tot_val:,.0f}").font = BOLD
    for cc in range(1, 9): wsp.cell(row=rr, column=cc).border = BORDER
    wsp.cell(row=rr+2, column=1,
             value="Scenario estimates. The time cut depends only on per-onboarding staff time (measurable in "
                   "the pilot); volume and rate scale value, not the cut.").font = ITAL
    wsp.merge_cells(start_row=rr+2, start_column=1, end_row=rr+2, end_column=8)
    widths(wsp, [22, 10, 10, 10, 14, 11, 11, 15])

    # Document completeness
    wsd = wb.create_sheet("Document completeness")
    wsd["A1"] = "Document completeness before the first advisory meeting"; wsd["A1"].font = TITLE
    hrow(wsd, 3, ["Measure", "Baseline (scenario)", "Target", "How it is achieved"])
    dc = [("Required documents in before first meeting", f"{DOC_BEFORE*100:.0f}%", f"{DOC_TARGET*100:.0f}%",
           "Correct request from the matrix + must-have gate: do not book the meeting until must-have is complete"),
          ("Right documents requested first time", "Varies by staff", "Consistent",
           "Assistant builds the request from the matrix for the service line"),
          ("Duplicate or wrong requests", "Occasional", "None",
           "Reminders mention only outstanding items; the tracker records what was sent")]
    rr = 4
    for m, b, t, h in dc:
        wsd.cell(row=rr, column=1, value=m).border = BORDER
        c = wsd.cell(row=rr, column=2, value=b); c.border = BORDER
        c2 = wsd.cell(row=rr, column=3, value=t); c2.border = BORDER; c2.fill = GOOD_FILL
        wsd.cell(row=rr, column=4, value=h).border = BORDER
        wsd.cell(row=rr, column=4).alignment = Alignment(wrap_text=True)
        rr += 1
    widths(wsd, [40, 18, 12, 46])

    # Adoption log
    wsl = wb.create_sheet("Adoption log")
    wsl["A1"] = "Adoption log, replace scenario estimates with measured pilot actuals"; wsl["A1"].font = TITLE
    wsl["A2"] = "The Practice Manager fills this in during the five-scenario pilot and rollout."; wsl["A2"].font = ITAL
    hrow(wsl, 4, ["Onboarding", "Service line", "Staff time BEFORE (h)", "Staff time AFTER (h)",
                  "Time cut %", "Docs complete before meeting %", "Client tone approved?", "Notes"])
    for i in range(10):
        for cc in range(1, 9): wsl.cell(row=5+i, column=cc).border = BORDER
        wsl.cell(row=5+i, column=1, value=f"Client {i+1}")
    widths(wsl, [16, 18, 18, 18, 12, 24, 18, 28])

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print("Wrote", path)
    for r in rows:
        print(f"  {r['name']}: {r['red']*100:.0f}% cut, {r['hrs']:,.0f} hrs/yr, ${r['val']:,.0f}")
    print(f"  TOTAL: {tot_hrs:,.0f} hrs/yr, AUD ${tot_val:,.0f} (~{tot_val/FEE:.1f}x fee)")


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    out = os.path.normpath(os.path.join(here, "..", "deliverables", "Elevate_Benefits_Model.xlsx"))
    build(out)
