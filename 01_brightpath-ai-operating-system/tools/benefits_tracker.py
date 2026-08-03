#!/usr/bin/env python3
"""
benefits_tracker.py, builds the BrightPath benefits-tracker spreadsheet from the
time-savings model in docs/before_after_comparison.md.

Illustrative model. All figures are scenario estimates under the assumptions below,
not measured client outcomes. Edit the ASSUMPTIONS block and re-run
to see the effect; the workbook rebuilds, including a live opportunity-matrix sheet and an
adoption-log template for recording real measured results later.

Usage:
    pip install openpyxl
    python tools/benefits_tracker.py         # -> deliverables/BrightPath_Benefits_Tracker.xlsx
"""

import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl")

# ----------------------------------------------------------------------------
# ASSUMPTIONS  (edit these, everything else recomputes)
# ----------------------------------------------------------------------------
WEEKS_PER_YEAR = 46

ADVICE = dict(
    before_hrs=5.0, after_hrs=3.3,   # drafting time per advice document
    docs_per_week=8, blended_rate=120,
)
COMMS = dict(
    before_min=6.0, after_min=4.2,   # handle time per composed reply
    staff=16, emails_per_day=15, days_per_week=5, share_ai=0.40, blended_rate=90,
)

# Opportunity matrix (value x feasibility), from docs/02_ai_opportunity_matrix.md
OPPORTUNITIES = [
    ("Advice preparation", 5, 4),
    ("Client communications", 4, 5),
    ("Meeting capture & actions", 4, 4),
    ("Proposals & engagement letters", 3, 4),
    ("Internal knowledge & search", 4, 3),
    ("Marketing content", 3, 3),
]

# ----------------------------------------------------------------------------
# Styling
# ----------------------------------------------------------------------------
NAVY, TEAL = "1F2A44", "2E7D8A"
HFILL = PatternFill("solid", fgColor=NAVY)
SUBFILL = PatternFill("solid", fgColor=TEAL)
INPUT_FILL = PatternFill("solid", fgColor="F5EDE0")
GOOD_FILL = PatternFill("solid", fgColor="E6F0EA")
HFONT = Font(color="FFFFFF", bold=True)
TITLE = Font(color=NAVY, bold=True, size=16)
BOLD = Font(bold=True)
ITAL = Font(italic=True, color="666666")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hrow(ws, row, headers, start=1, fill=HFILL):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=h)
        c.fill = fill
        c.font = HFONT
        c.border = BORDER


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ----------------------------------------------------------------------------
# Compute
# ----------------------------------------------------------------------------
def compute():
    a = ADVICE
    a_saving = a["before_hrs"] - a["after_hrs"]
    a_red = a_saving / a["before_hrs"]
    a_hrs_yr = a["docs_per_week"] * a_saving * WEEKS_PER_YEAR
    a_val = a_hrs_yr * a["blended_rate"]

    c = COMMS
    c_saving_min = c["before_min"] - c["after_min"]
    c_red = c_saving_min / c["before_min"]
    c_elig_wk = c["staff"] * c["emails_per_day"] * c["days_per_week"] * c["share_ai"]
    c_hrs_wk = c_elig_wk * c_saving_min / 60
    c_hrs_yr = c_hrs_wk * WEEKS_PER_YEAR
    c_val = c_hrs_yr * c["blended_rate"]

    return dict(
        a_saving=a_saving, a_red=a_red, a_hrs_yr=a_hrs_yr, a_val=a_val,
        c_saving_min=c_saving_min, c_red=c_red, c_elig_wk=c_elig_wk,
        c_hrs_yr=c_hrs_yr, c_val=c_val,
        tot_hrs=a_hrs_yr + c_hrs_yr, tot_val=a_val + c_val,
    )


# ----------------------------------------------------------------------------
# Build workbook
# ----------------------------------------------------------------------------
def build(path):
    r = compute()
    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "BrightPath AI Operating System, Benefits Tracker"
    ws["A1"].font = TITLE
    ws["A2"] = "Illustrative model, figures are scenario estimates under the assumptions shown."
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A3"] = "Edit the Assumptions sheet and re-run tools/benefits_tracker.py to update."
    ws["A3"].font = ITAL

    tiles = [
        ("Two-pilot hours reclaimed / yr", f"{r['tot_hrs']:,.0f} hrs"),
        ("Indicative value / yr (scenario)", f"AUD ${r['tot_val']:,.0f}"),
        ("Advice cycle-time reduction", f"{r['a_red']*100:,.0f}%"),
        ("Client-comms handle-time reduction", f"{r['c_red']*100:,.0f}%"),
    ]
    row = 5
    for i, (label, val) in enumerate(tiles):
        col = 1 + i * 2
        lc = ws.cell(row=row, column=col, value=label)
        lc.fill = SUBFILL
        lc.font = HFONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        vc = ws.cell(row=row + 1, column=col, value=val)
        vc.font = Font(size=15, bold=True, color=NAVY)
        vc.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 20
        ws.column_dimensions[get_column_letter(col + 1)].width = 3

    ws["A9"] = ("Budget benchmark: engagement fee AUD $18,000. Modelled first-year value of the "
                "two pilots alone ≈ AUD ${:,.0f} (~{:.0f}× the fee) on the stated scenario assumptions. "
                "Primary metric is cycle-time %, which is robust to the volume/$ assumptions.").format(
        r["tot_val"], r["tot_val"] / 18000)
    ws["A9"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A9:H12")

    # ---- Assumptions ----
    wsa = wb.create_sheet("Assumptions")
    wsa["A1"] = "Assumptions (editable inputs)"
    wsa["A1"].font = TITLE
    wsa["A2"] = "Highlighted cells are the inputs. Change them, re-run the script."
    wsa["A2"].font = ITAL
    hrow(wsa, 4, ["Parameter", "Value", "Unit / note"])
    rows = [
        ("Working weeks per year", WEEKS_PER_YEAR, "weeks"),
        (", Advice preparation,", "", ""),
        ("Drafting time BEFORE", ADVICE["before_hrs"], "hrs / advice document"),
        ("Drafting time AFTER", ADVICE["after_hrs"], "hrs / advice document"),
        ("Advice documents per week", ADVICE["docs_per_week"], "documents"),
        ("Blended internal cost (advice)", ADVICE["blended_rate"], "AUD / hour"),
        (", Client communications,", "", ""),
        ("Handle time BEFORE", COMMS["before_min"], "min / composed reply"),
        ("Handle time AFTER", COMMS["after_min"], "min / composed reply"),
        ("Staff doing client email", COMMS["staff"], "people"),
        ("Composed replies per person / day", COMMS["emails_per_day"], "emails"),
        ("Working days per week", COMMS["days_per_week"], "days"),
        ("Share of emails suited to AI assist", COMMS["share_ai"], "fraction"),
        ("Blended internal cost (client services)", COMMS["blended_rate"], "AUD / hour"),
    ]
    rr = 5
    for name, val, unit in rows:
        wsa.cell(row=rr, column=1, value=name)
        vc = wsa.cell(row=rr, column=2, value=val)
        wsa.cell(row=rr, column=3, value=unit)
        if name.startswith(", "):
            wsa.cell(row=rr, column=1).font = BOLD
        elif val != "":
            vc.fill = INPUT_FILL
            vc.border = BORDER
        rr += 1
    widths(wsa, [38, 14, 26])

    # ---- Pilot models ----
    wsp = wb.create_sheet("Pilot models")
    wsp["A1"] = "Pilot models, computed from Assumptions"
    wsp["A1"].font = TITLE
    hrow(wsp, 3, ["Metric", "Advice preparation", "Client communications"])
    model_rows = [
        ("Time saved per item", f"{r['a_saving']:.1f} hrs / doc", f"{r['c_saving_min']:.1f} min / email"),
        ("Cycle-time reduction", f"{r['a_red']*100:.0f}%", f"{r['c_red']*100:.0f}%"),
        ("Meets ≥20% target", "Yes" if r['a_red'] >= .2 else "No", "Yes" if r['c_red'] >= .2 else "No"),
        ("Eligible volume / week",
         f"{ADVICE['docs_per_week']} docs", f"{r['c_elig_wk']:.0f} emails"),
        ("Hours reclaimed / year", f"{r['a_hrs_yr']:,.0f}", f"{r['c_hrs_yr']:,.0f}"),
        ("Indicative value / year (AUD)", f"${r['a_val']:,.0f}", f"${r['c_val']:,.0f}"),
    ]
    rr = 4
    for m, av, cv in model_rows:
        wsp.cell(row=rr, column=1, value=m).font = BOLD
        wsp.cell(row=rr, column=2, value=av)
        wsp.cell(row=rr, column=3, value=cv)
        for cc in range(1, 4):
            wsp.cell(row=rr, column=cc).border = BORDER
        if m == "Meets ≥20% target":
            wsp.cell(row=rr, column=2).fill = GOOD_FILL
            wsp.cell(row=rr, column=3).fill = GOOD_FILL
        rr += 1
    wsp.cell(row=rr + 1, column=1,
             value="Scenario estimates. Cycle-time % depends only on per-item time (directly measurable); "
                   "volume and $ assumptions scale value, not the %.").font = ITAL
    wsp.merge_cells(start_row=rr + 1, start_column=1, end_row=rr + 1, end_column=3)
    widths(wsp, [30, 24, 24])

    # ---- Opportunity matrix ----
    wsm = wb.create_sheet("Opportunity matrix")
    wsm["A1"] = "AI Opportunity Matrix (Priority = Value × Feasibility)"
    wsm["A1"].font = TITLE
    hrow(wsm, 3, ["Workflow", "Value (1-5)", "Feasibility (1-5)", "Priority", "Rank"])
    ranked = sorted(OPPORTUNITIES, key=lambda x: -(x[1] * x[2]))
    rr = 4
    for i, (name, v, f) in enumerate(sorted(OPPORTUNITIES, key=lambda x: -(x[1]*x[2]))):
        pri = v * f
        rank = i + 1
        wsm.cell(row=rr, column=1, value=name)
        wsm.cell(row=rr, column=2, value=v)
        wsm.cell(row=rr, column=3, value=f)
        wsm.cell(row=rr, column=4, value=pri)
        wsm.cell(row=rr, column=5, value=("Pilot" if rank <= 2 else rank))
        for cc in range(1, 6):
            wsm.cell(row=rr, column=cc).border = BORDER
        if rank <= 2:
            for cc in range(1, 6):
                wsm.cell(row=rr, column=cc).fill = GOOD_FILL
        rr += 1
    widths(wsm, [32, 12, 15, 10, 10])

    # ---- Adoption log (template for real actuals) ----
    wsl = wb.create_sheet("Adoption log")
    wsl["A1"] = "Adoption log, replace scenario estimates with measured actuals"
    wsl["A1"].font = TITLE
    wsl["A2"] = "AI Champions fill this in during the pilot and 90-day adoption. This is where scenario becomes real."
    wsl["A2"].font = ITAL
    hrow(wsl, 4, ["Week", "Workflow", "Items measured", "Avg time BEFORE",
                  "Avg time AFTER", "Actual reduction %", "Notes / guardrails"])
    for i in range(12):
        for cc in range(1, 8):
            wsl.cell(row=5 + i, column=cc).border = BORDER
        wsl.cell(row=5 + i, column=1, value=i + 1)
    widths(wsl, [8, 22, 14, 16, 16, 16, 34])

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print("Wrote", path)
    print(f"  Advice: {r['a_red']*100:.0f}% reduction, {r['a_hrs_yr']:,.0f} hrs/yr, ${r['a_val']:,.0f}")
    print(f"  Comms:  {r['c_red']*100:.0f}% reduction, {r['c_hrs_yr']:,.0f} hrs/yr, ${r['c_val']:,.0f}")
    print(f"  Total:  {r['tot_hrs']:,.0f} hrs/yr, AUD ${r['tot_val']:,.0f} (~{r['tot_val']/18000:.0f}x fee)")


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(here, "..", "deliverables", "BrightPath_Benefits_Tracker.xlsx")
    build(os.path.normpath(out))
