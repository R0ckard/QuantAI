#!/usr/bin/env python3
"""
benefits_tracker.py, builds the Vertex Engineering benefits-tracker spreadsheet
from the time-savings model in docs/before_after_comparison.md.

Scenario model. All figures are estimates under the assumptions below, not
measured client outcomes. Edit the ASSUMPTIONS block and re-run to update the
workbook, which includes a live ten-workflow prioritisation-matrix sheet and an
adoption-log template for recording real pilot measurements.

Usage:
    pip install openpyxl
    python tools/benefits_tracker.py    # -> deliverables/Vertex_Benefits_Tracker.xlsx
"""

import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl")

# ----------------------------------------------------------------------------
# ASSUMPTIONS  (edit these, everything else recomputes)
# ----------------------------------------------------------------------------
STAFF = 52
WEEKS_PER_YEAR = 46
RATE = 95            # AUD/hour blended professional cost of time
FEE = 35000          # AUD engagement fee benchmark
SUCCESS_BAR = 0.15   # brief: >= 15% cycle-time improvement

# Pilots are modelled in HOURS per item (these are professional, hours-scale tasks).
# The saving is only on the drafting / formatting / QA / admin portion.
PILOTS = [
    # name, before_hours, after_hours, volume_per_year
    ("Proposals and bids",                15.0, 11.0, 124),
    ("Technical report drafting and QA",  12.0,  8.4, 191),
    ("Project status reporting",           1.5,  1.0, 460),
]

# Ten-workflow prioritisation (Value, Feasibility, Risk). Priority = V * F / Risk.
# Matches docs/03_workflow_prioritisation_matrix.md.
WORKFLOWS = [
    # name, value, feasibility, risk, decision
    ("Proposals and bids",                             5, 4, 2, "Pilot"),
    ("Technical report drafting and QA",               5, 3, 3, "Pilot"),
    ("Project status reporting",                       4, 5, 1, "Pilot"),
    ("Capability statements and CVs for bids",         4, 5, 2, "Fast follow"),
    ("Meeting notes and actions",                      4, 5, 2, "Fast follow"),
    ("Knowledge retrieval and reuse",                  4, 3, 2, "Phase 2"),
    ("Client and internal comms drafting",             3, 5, 2, "Fast follow"),
    ("Finance management-reporting narratives",        3, 4, 2, "Phase 2"),
    ("Recruitment (job ads, interview prep)",          3, 4, 3, "Phase 2"),
    ("Internal policy and process documentation",      3, 4, 1, "Fast follow"),
]

# ----------------------------------------------------------------------------
# Styling (QuantAI brand)
# ----------------------------------------------------------------------------
NAVY, TEAL = "0F1B2A", "14655C"
HFILL = PatternFill("solid", fgColor=NAVY)
SUBFILL = PatternFill("solid", fgColor=TEAL)
INPUT_FILL = PatternFill("solid", fgColor="F5EDE0")
GOOD_FILL = PatternFill("solid", fgColor="E6F0EA")
EXCL_FILL = PatternFill("solid", fgColor="F5EDE0")
HFONT = Font(color="FFFFFF", bold=True)
TITLE = Font(name="Cambria", color=NAVY, bold=True, size=16)
BOLD = Font(bold=True)
ITAL = Font(italic=True, color="666666")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hrow(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=h)
        c.fill = HFILL; c.font = HFONT; c.border = BORDER


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def compute():
    rows = []
    tot_hrs = tot_val = 0
    for name, before, after, vol in PILOTS:
        saved = before - after
        red = saved / before
        hrs_yr = vol * saved
        val = hrs_yr * RATE
        tot_hrs += hrs_yr; tot_val += val
        rows.append(dict(name=name, before=before, after=after, saved=saved,
                         red=red, vol=vol, hrs=hrs_yr, val=val))
    return rows, tot_hrs, tot_val


def build(path):
    rows, tot_hrs, tot_val = compute()
    wb = Workbook()

    # Summary
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "Vertex Engineering AI Operating Model, Benefits Tracker"; ws["A1"].font = TITLE
    ws["A2"] = "Scenario model, figures are estimates under the assumptions shown, not measured outcomes."
    ws["A2"].font = ITAL
    tiles = [("Three-pilot hours / yr", f"{tot_hrs:,.0f} hrs"),
             ("Indicative value / yr", f"AUD ${tot_val:,.0f}"),
             ("Value vs fee", f"~{tot_val/FEE:.1f}x"),
             ("Min. cycle-time cut", f"{min(r['red'] for r in rows)*100:.0f}%")]
    row = 4
    for i, (label, val) in enumerate(tiles):
        col = 1 + i*2
        c = ws.cell(row=row, column=col, value=label); c.fill = SUBFILL; c.font = HFONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        v = ws.cell(row=row+1, column=col, value=val); v.font = Font(size=14, bold=True, color=NAVY)
        v.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions[get_column_letter(col+1)].width = 3
    ws["A8"] = ("Brief success bar: at least 15% cycle-time improvement on each pilot without unacceptable "
                "quality loss. All three pilots clear it (27-33% on the scenario assumptions). Engagement fee "
                "benchmark AUD ${:,.0f}; modelled first-year value of the three pilots is about AUD ${:,.0f} "
                "(~{:.1f}x the fee). Primary metric is cycle-time %, which depends only on per-task time and is "
                "robust to the volume and rate assumptions; the dollar figure sizes the prize, it is not the "
                "headline.").format(FEE, tot_val, tot_val/FEE)
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="top"); ws.merge_cells("A8:H12")

    # Assumptions
    wsa = wb.create_sheet("Assumptions")
    wsa["A1"] = "Assumptions (editable inputs)"; wsa["A1"].font = TITLE
    wsa["A2"] = "Highlighted cells are inputs. Change them, re-run the script."; wsa["A2"].font = ITAL
    hrow(wsa, 4, ["Parameter", "Value", "Unit / note"])
    base = [("Staff", STAFF, "people"),
            ("Working weeks / year", WEEKS_PER_YEAR, "weeks"),
            ("Blended professional cost", RATE, "AUD / hour"),
            ("Engagement fee benchmark", FEE, "AUD"),
            ("Success bar", f"{SUCCESS_BAR*100:.0f}%", "min cycle-time cut per pilot")]
    rr = 5
    for n, v, u in base:
        wsa.cell(row=rr, column=1, value=n)
        c = wsa.cell(row=rr, column=2, value=v); c.fill = INPUT_FILL; c.border = BORDER
        wsa.cell(row=rr, column=3, value=u); rr += 1
    rr += 1
    wsa.cell(row=rr, column=1, value="Per-pilot inputs (hours per item)").font = BOLD; rr += 1
    hrow(wsa, rr, ["Pilot", "Before (h)", "After (h)", "Volume / yr"]); rr += 1
    for name, before, after, vol in PILOTS:
        wsa.cell(row=rr, column=1, value=name)
        for col, val in [(2, before), (3, after), (4, vol)]:
            c = wsa.cell(row=rr, column=col, value=val); c.fill = INPUT_FILL; c.border = BORDER
        rr += 1
    wsa.cell(row=rr+1, column=1,
             value="Saving is only on the drafting / formatting / QA / admin portion. On technical reports "
                   "no time is saved on, or taken from, the engineering work itself.").font = ITAL
    wsa.merge_cells(start_row=rr+1, start_column=1, end_row=rr+1, end_column=4)
    widths(wsa, [34, 14, 14, 16])

    # Pilot models
    wsp = wb.create_sheet("Pilot models")
    wsp["A1"] = "Pilot models, computed from Assumptions"; wsp["A1"].font = TITLE
    hrow(wsp, 3, ["Pilot", "Before", "After", "Cycle-time cut", ">= 15% bar",
                  "Volume/yr", "Hours/yr", "Value/yr (AUD)"])
    rr = 4
    for r in rows:
        wsp.cell(row=rr, column=1, value=r["name"])
        wsp.cell(row=rr, column=2, value=f'{r["before"]:.1f} h')
        wsp.cell(row=rr, column=3, value=f'{r["after"]:.1f} h')
        wsp.cell(row=rr, column=4, value=f'{r["red"]*100:.0f}%')
        tgt = wsp.cell(row=rr, column=5, value="Yes" if r["red"] >= SUCCESS_BAR else "No")
        if r["red"] >= SUCCESS_BAR: tgt.fill = GOOD_FILL
        wsp.cell(row=rr, column=6, value=r["vol"])
        wsp.cell(row=rr, column=7, value=round(r["hrs"]))
        wsp.cell(row=rr, column=8, value=f'${r["val"]:,.0f}')
        for cc in range(1, 9): wsp.cell(row=rr, column=cc).border = BORDER
        rr += 1
    wsp.cell(row=rr, column=1, value="TOTAL").font = BOLD
    wsp.cell(row=rr, column=7, value=round(tot_hrs)).font = BOLD
    wsp.cell(row=rr, column=8, value=f"${tot_val:,.0f}").font = BOLD
    for cc in range(1, 9): wsp.cell(row=rr, column=cc).border = BORDER
    wsp.cell(row=rr+2, column=1,
             value="Scenario estimates. Cycle-time % depends only on per-task time (measurable in the pilot); "
                   "volume and rate scale value, not the %.").font = ITAL
    wsp.merge_cells(start_row=rr+2, start_column=1, end_row=rr+2, end_column=8)
    widths(wsp, [30, 10, 10, 14, 12, 11, 11, 15])

    # Prioritisation matrix
    wsm = wb.create_sheet("Prioritisation matrix")
    wsm["A1"] = "Ten-workflow prioritisation (Priority = Value x Feasibility / Risk)"
    wsm["A1"].font = TITLE
    wsm["A2"] = "Scores 1-5. Weights are visible so the governance forum can re-weight them."; wsm["A2"].font = ITAL
    hrow(wsm, 4, ["Workflow", "Value", "Feasibility", "Risk", "Priority", "Decision"])
    scored = []
    for name, v, f, risk, dec in WORKFLOWS:
        pri = round(v * f / risk)
        scored.append((name, v, f, risk, pri, dec))
    scored.sort(key=lambda x: -x[4])
    rr = 5
    for name, v, f, risk, pri, dec in scored:
        for col, val in [(1, name), (2, v), (3, f), (4, risk), (5, pri), (6, dec)]:
            c = wsm.cell(row=rr, column=col, value=val); c.border = BORDER
        if dec == "Pilot":
            for cc in range(1, 7): wsm.cell(row=rr, column=cc).fill = GOOD_FILL
        rr += 1
    # Excluded row
    for col, val in [(1, "Engineering design, calculation, verification, certification"),
                     (2, "-"), (3, "-"), (4, 5), (5, "Excluded"), (6, "Out of bounds")]:
        c = wsm.cell(row=rr, column=col, value=val); c.border = BORDER; c.fill = EXCL_FILL
    wsm.cell(row=rr+2, column=1,
             value="Engineering design, calculation, verification and certification are excluded by design. "
                   "AI never performs, checks or certifies engineering work; a qualified engineer owns it.").font = ITAL
    wsm.merge_cells(start_row=rr+2, start_column=1, end_row=rr+2, end_column=6)
    widths(wsm, [50, 10, 12, 8, 10, 14])

    # Adoption log
    wsl = wb.create_sheet("Adoption log")
    wsl["A1"] = "Adoption log, replace scenario estimates with measured pilot actuals"; wsl["A1"].font = TITLE
    wsl["A2"] = "AI Champions fill this in during the pilot and 90-day rollout."; wsl["A2"].font = ITAL
    hrow(wsl, 4, ["Week", "Workflow", "Items measured", "Avg BEFORE (h)",
                  "Avg AFTER (h)", "Actual cut %", "Quality pass %", "Notes"])
    for i in range(12):
        for cc in range(1, 9): wsl.cell(row=5+i, column=cc).border = BORDER
        wsl.cell(row=5+i, column=1, value=i+1)
    widths(wsl, [8, 30, 14, 15, 15, 14, 16, 30])

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print("Wrote", path)
    for r in rows:
        print(f"  {r['name']}: {r['red']*100:.0f}% cut, {r['hrs']:,.0f} hrs/yr, ${r['val']:,.0f}")
    print(f"  TOTAL: {tot_hrs:,.0f} hrs/yr, AUD ${tot_val:,.0f} (~{tot_val/FEE:.1f}x fee)")


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    out = os.path.normpath(os.path.join(here, "..", "deliverables", "Vertex_Benefits_Tracker.xlsx"))
    build(out)
